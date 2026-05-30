"""
导出模块 - 重构版 v2.0
仅保留 Excel 明细表和合并 PDF 导出功能
移除问题发票导出相关代码
"""
import logging
import shutil
from pathlib import Path
from typing import List, Optional
from datetime import datetime

import fitz

logger = logging.getLogger(__name__)


def export_excel(invoices: List[dict], output_path: Path) -> Path:
    """
    生成报销 Excel 明细表
    列：序号、开票日期、发票号码、发票代码、项目名称、规格型号、
        购买方名称、购买方税号、销售方名称、销售方税号、
        税率、税额、价税合计、分类、归属项目、归属人、备注
    按开票日期升序排列，最后一行合计
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "发票明细"

    # 标题行
    headers = [
        "序号", "开票日期", "发票号码", "发票代码", "项目名称", "规格型号",
        "购买方名称", "购买方税号", "销售方名称", "销售方税号",
        "税率", "税额", "价税合计", "分类", "归属项目", "归属人", "备注"
    ]
    col_widths = [6, 14, 22, 14, 30, 20, 30, 22, 30, 22, 8, 14, 16, 10, 16, 12, 20]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # 数据行（按日期排序）
    sorted_inv = sorted(invoices, key=lambda x: x.get("invoice_date") or "")
    even_fill = PatternFill("solid", fgColor="DCE6F1")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    total_amount = 0.0

    for row_idx, inv in enumerate(sorted_inv, 2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        amount = inv.get("amount_with_tax") or 0
        total_amount += amount

        # 税率格式化
        tax_rate = inv.get("tax_rate")
        tax_rate_str = f"{tax_rate * 100:.0f}%" if tax_rate is not None else ""

        values = [
            row_idx - 1,
            inv.get("invoice_date") or "",
            inv.get("invoice_number") or "",
            inv.get("invoice_code") or "",
            inv.get("commodity_name") or "",
            inv.get("specification_model") or "",
            inv.get("buyer_name") or "",
            inv.get("buyer_tax_num") or "",
            inv.get("seller_name") or "",
            inv.get("seller_tax_num") or "",
            tax_rate_str,
            inv.get("tax_amount") or 0,
            amount,
            inv.get("category") or "",
            inv.get("belong_project") or "",
            inv.get("belong_person") or "",
            inv.get("remark") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in (12, 13):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col == 11:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # 合计行
    total_row = len(sorted_inv) + 2
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    total_font = Font(name="微软雅黑", bold=True, size=11)

    ws.merge_cells(f"A{total_row}:L{total_row}")
    total_label = ws.cell(row=total_row, column=1, value=f"合计（共 {len(sorted_inv)} 张）")
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    total_label.border = border

    total_cell = ws.cell(row=total_row, column=13, value=total_amount)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = border
    ws.row_dimensions[total_row].height = 22

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    logger.info(f"Excel 导出完成: {output_path}，共 {len(sorted_inv)} 张，合计 {total_amount:.2f}")
    return output_path


def _insert_file_as_pdf_page(merged: fitz.Document, file_path: Path) -> bool:
    """
    将 PDF 或图片文件插入到合并文档中。
    返回 True 表示成功。
    """
    # 方法1：尝试作为 PDF 打开
    try:
        doc = fitz.open(str(file_path))
        # 确认是可读的 PDF（有页面）
        if doc.page_count > 0:
            merged.insert_pdf(doc)
            doc.close()
            return True
        doc.close()
    except Exception:
        pass

    # 方法2：尝试作为图片打开（PNG/JPG/BMP 等）
    try:
        pix = fitz.Pixmap(str(file_path))
        # 创建一个与图片等尺寸的页面
        page = merged.new_page(width=pix.width, height=pix.height)
        page.insert_image(page.rect, pixmap=pix)
        return True
    except Exception as e:
        logger.warning(f"无法读取文件（非 PDF/图片）: {file_path.name} - {e}")
        return False


def export_merged_pdf(invoices: List[dict], output_path: Path,
                      attachments_map: Optional[dict] = None) -> Optional[Path]:
    """
    合并发票为单一 PDF。
    若 attachments_map 提供（{invoice_id: [attachment_records]}），
    则在每张发票页后紧跟其附件图片页。
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    merged = fitz.open()
    count = 0
    for inv in sorted(invoices, key=lambda x: x.get("invoice_date") or ""):
        src = Path(inv.get("stored_path") or "")
        if not src.exists():
            logger.warning(f"文件不存在，跳过: {src}")
            continue

        # 插入发票页
        if _insert_file_as_pdf_page(merged, src):
            count += 1
        else:
            logger.warning(f"合并跳过（无法识别格式）: {src.name}")
            continue

        # 插入该发票的附件页（紧跟在发票后面）
        if attachments_map and inv["id"] in attachments_map:
            for att in attachments_map[inv["id"]]:
                att_path = Path(att["stored_path"])
                if not att_path.exists():
                    continue
                # 仅图片附件有视觉意义，PDF 附件跳过（避免叠加整本 PDF）
                if att_path.suffix.lower() in (".png", ".jpg", ".jpeg", ".bmp", ".webp", ".gif"):
                    _insert_file_as_pdf_page(merged, att_path)

    if count == 0:
        logger.warning("没有可合并的文件")
        return None

    merged.save(str(output_path))
    merged.close()
    logger.info(f"合并 PDF 完成: {output_path}，共 {count} 张发票"
                f"{' + 附件' if attachments_map else ''}")
    return output_path


def build_export_label(filters: dict) -> str:
    """根据筛选条件生成导出文件名标签"""
    parts = []
    if filters.get("date_from") and filters.get("date_to"):
        parts.append(f"{filters['date_from']}至{filters['date_to']}")
    elif filters.get("date_from"):
        parts.append(f"{filters['date_from']}起")
    elif filters.get("date_to"):
        parts.append(f"至{filters['date_to']}")
    if filters.get("buyer"):
        parts.append(f"购买方{filters['buyer']}")
    if filters.get("seller"):
        parts.append(f"销售方{filters['seller']}")
    if filters.get("project"):
        parts.append(f"项目{filters['project']}")
    if filters.get("person"):
        parts.append(f"归属{filters['person']}")
    if not parts:
        parts.append(datetime.now().strftime("%Y%m%d"))
    return "_".join(parts)